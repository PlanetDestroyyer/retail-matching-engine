"""
step3_classifier.py — Case A / B / C classifier.
Reads match_diffs.jsonl → writes comparison_results.jsonl + comparison_results.xlsx

Usage:
    uv run python src/step3_classifier.py --limit 5
"""

from __future__ import annotations

import sys
import argparse
import json
from pathlib import Path
from datetime import datetime, timezone

sys.path.insert(0, str(Path(__file__).parent))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from rich.console import Console
from rich.table import Table

from utils import get_logger, jsonl_read, jsonl_append, load_pairs

logger = get_logger("step3_classifier")
console = Console()

DIFFS_IN    = Path("data/output/match_diffs.jsonl")
RESULTS_OUT = Path("data/output/comparison_results.jsonl")
EXCEL_OUT   = Path("data/output/comparison_results.xlsx")

# Keywords that classify a mismatched key as size-type or color-type
SIZE_KEYWORDS  = {"size", "dimension", "length", "width", "height", "weight",
                  "volume", "capacity", "fit", "oz", "ml", "kg", "cm",
                  "measurement", "inseam", "waist", "chest", "neck", "sleeve"}
COLOR_KEYWORDS = {"color", "colour", "shade", "finish", "tone", "hue", "tint"}

# Result labels
EXACT_MATCH  = "CASE_A_EXACT_MATCH"
SIZE_DIFF    = "CASE_B_SIZE_DIFF"
COLOR_DIFF   = "CASE_C_COLOR_DIFF"
NO_MATCH     = "NO_MATCH"

import re

def clean_brand(b: str) -> str:
    b = b.lower().strip()
    b = re.sub(r"[^\w\s]", "", b)  # remove punctuation
    # Remove common suffixes/prefixes
    b = re.sub(r"\b(co|corp|inc|llc|ltd|company|brand|premium|classic|basics?)\b", "", b)
    return " ".join(b.split())


def classify_key(key: str) -> str:
    """Returns 'size', 'color', or 'other' based on keyword presence in the key name."""
    k = key.lower()
    if any(kw in k for kw in SIZE_KEYWORDS):
        return "size"
    if any(kw in k for kw in COLOR_KEYWORDS):
        return "color"
    return "other"


def classify(diff: dict) -> str:
    """
    Apply Case A / B / C logic to a diff record.

    Case A: No mismatches → EXACT MATCH
    Case B: All mismatches are size-type keys → SIZE DIFF
    Case C: All mismatches are color-type keys → COLOR DIFF
    else  → NO MATCH
    """
    # 1. Validation checks
    t1 = diff.get("title_link_1", "").lower()
    t2 = diff.get("title_link_2", "").lower()
    if not t1 or not t2 or "robot or human" in t1 or "robot or human" in t2:
        return NO_MATCH
    if "blocked" in t1 or "blocked" in t2 or "captcha" in t1 or "captcha" in t2:
        return NO_MATCH

    # 2. Brand alignment check
    feature_diff = diff.get("feature_diff", {})
    b1 = diff.get("brand_link_1", "")
    b2 = diff.get("brand_link_2", "")
    if not b1 or not b2:
        for k, d in feature_diff.items():
            if k in ("brand", "brand_name"):
                b1 = b1 or d.get("link_1_value")
                b2 = b2 or d.get("link_2_value")
                break

    if b1 and b2:
        cb1 = clean_brand(str(b1))
        cb2 = clean_brand(str(b2))
        if cb1 and cb2:
            if cb1 != cb2 and cb1 not in cb2 and cb2 not in cb1:
                return NO_MATCH
    elif b1:
        cb1 = clean_brand(str(b1))
        if cb1 and cb1 not in t2:
            w1 = t1.split()[0] if t1.split() else ""
            w2 = t2.split()[0] if t2.split() else ""
            if w1 and w2 and w1 != w2 and w1 not in ("men's", "mens", "women's", "womens", "unisex", "kids", "boys", "girls"):
                return NO_MATCH
    elif b2:
        cb2 = clean_brand(str(b2))
        if cb2 and cb2 not in t1:
            w1 = t1.split()[0] if t1.split() else ""
            w2 = t2.split()[0] if t2.split() else ""
            if w1 and w2 and w1 != w2 and w2 not in ("men's", "mens", "women's", "womens", "unisex", "kids", "boys", "girls"):
                return NO_MATCH

    # 3. Title similarity check
    title_sim = diff.get("title_similarity_pct", 0)
    if title_sim < 40.0:
        return NO_MATCH

    # 4. Check aligned mismatches only (keys present on both sides but with different values)
    aligned_mismatches = []
    for k, d in feature_diff.items():
        if d.get("link_1_key") is not None and d.get("link_2_key") is not None:
            if not d.get("match", True):
                aligned_mismatches.append(k)

    if not aligned_mismatches:
        if diff.get("total_features_compared", 0) == 0:
            if title_sim < 85.0:
                return NO_MATCH
        return EXACT_MATCH

    key_types = {classify_key(k) for k in aligned_mismatches}

    if key_types == {"size"}:
        return SIZE_DIFF

    if key_types == {"color"}:
        return COLOR_DIFF

    return NO_MATCH


def build_result_row(diff: dict, meta_map: dict, result: str) -> dict:
    """Combine diff + classification + CSV meta into output row."""
    pid = diff["pair_id"]
    meta = meta_map.get(pid, {})

    # Compact mismatched features for output
    mismatched_features = {
        k: {
            "link_1": v.get("link_1_value"),
            "link_2": v.get("link_2_value"),
        }
        for k, v in diff.get("feature_diff", {}).items()
        if not v.get("match", True)
    }

    matched_features = {
        k: v.get("link_1_value")
        for k, v in diff.get("feature_diff", {}).items()
        if v.get("match", False)
    }

    return {
        "pair_id": pid,
        "link_1_url": diff.get("link_1_url", ""),
        "link_2_url": diff.get("link_2_url", ""),
        "brand_link_1": diff.get("brand_link_1", ""),
        "brand_link_2": diff.get("brand_link_2", ""),
        "original_csv_label": meta.get("original_match_type", ""),
        "our_result": result,
        "title_link_1": diff.get("title_link_1", ""),
        "title_link_2": diff.get("title_link_2", ""),
        "title_similarity_pct": diff.get("title_similarity_pct", -1),
        "image_similarity_pct": diff.get("image_similarity_pct", -1),
        "total_features_compared": diff.get("total_features_compared", 0),
        "total_matched": diff.get("total_matched", 0),
        "total_mismatched": diff.get("total_mismatched", 0),
        "mismatched_features": mismatched_features,
        "matched_features": matched_features,
        "product_type": meta.get("product_type", ""),
        "department": meta.get("department", ""),
        "classified_at": datetime.now(timezone.utc).isoformat(),
    }


# ── Excel writer ──────────────────────────────────────────────────────────────

RESULT_COLORS = {
    EXACT_MATCH: "C6EFCE",  # Green
    SIZE_DIFF:   "FFEB9C",  # Amber
    COLOR_DIFF:  "BDD7EE",  # Blue
    NO_MATCH:    "FFC7CE",  # Red
}

RESULT_LABELS = {
    EXACT_MATCH: "✅ Exact Match",
    SIZE_DIFF:   "⚠️ Size Diff",
    COLOR_DIFF:  "⚠️ Color Diff",
    NO_MATCH:    "❌ No Match",
}


def write_excel(results: list[dict]):
    rows = []
    for r in results:
        rows.append({
            "Pair ID":            r["pair_id"],
            "Link 1 URL":         r["link_1_url"],
            "Link 2 URL":         r["link_2_url"],
            "Result":             RESULT_LABELS.get(r["our_result"], r["our_result"]),
            "CSV Label":          r["original_csv_label"],
            "Title (Link 1)":     r["title_link_1"],
            "Title (Link 2)":     r["title_link_2"],
            "Title Sim %":        r["title_similarity_pct"],
            "Image Sim %":        r["image_similarity_pct"],
            "Features Compared":  r["total_features_compared"],
            "Matched":            r["total_matched"],
            "Mismatched":         r["total_mismatched"],
            "Mismatched Features":json.dumps(r["mismatched_features"], ensure_ascii=False),
            "Matched Features":   json.dumps(r["matched_features"],    ensure_ascii=False),
            "Product Type":       r["product_type"],
        })

    df = pd.DataFrame(rows)
    df.to_excel(EXCEL_OUT, index=False, sheet_name="Comparison Results")

    wb = load_workbook(EXCEL_OUT)
    ws = wb.active

    # Header styling
    header_fill = PatternFill(fill_type="solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Result column (D = column 4) — color per row
    result_col_idx = 4
    for row_idx, r in enumerate(results, start=2):
        color = RESULT_COLORS.get(r["our_result"], "FFFFFF")
        fill  = PatternFill(fill_type="solid", fgColor=color)
        ws.cell(row=row_idx, column=result_col_idx).fill = fill

    # Auto column widths (capped at 60)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"
    wb.save(EXCEL_OUT)
    logger.info(f"Excel saved → {EXCEL_OUT}")


def run_classifier(pairs: list[dict]):
    pair_ids = {p["pair_id"] for p in pairs}
    meta_map = {p["pair_id"]: p.get("meta", {}) for p in pairs}

    diffs = [d for d in jsonl_read(DIFFS_IN) if d["pair_id"] in pair_ids]
    logger.info(f"Classifying {len(diffs)} pairs.")

    results = []
    with Progress(SpinnerColumn(), TextColumn("[bold orange]{task.description}"),
                  BarColumn(), TaskProgressColumn()) as progress:
        task = progress.add_task("Classifying...", total=len(diffs))

        for diff in diffs:
            result = classify(diff)
            row = build_result_row(diff, meta_map, result)
            jsonl_append(RESULTS_OUT, row)
            results.append(row)
            progress.advance(task)

    # Summary table
    from collections import Counter
    counts = Counter(r["our_result"] for r in results)
    table = Table(title="Classification Summary", show_header=True)
    table.add_column("Result", style="bold")
    table.add_column("Count",  justify="right")
    table.add_column("Pct",    justify="right")
    total = len(results)
    for label in [EXACT_MATCH, SIZE_DIFF, COLOR_DIFF, NO_MATCH]:
        n = counts.get(label, 0)
        table.add_row(RESULT_LABELS[label], str(n), f"{n/total*100:.1f}%" if total else "0%")
    console.print(table)

    write_excel(results)
    logger.info(f"Results → {RESULTS_OUT} | {EXCEL_OUT}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=None)
    args = parser.parse_args()
    pairs = load_pairs("data.csv", limit=args.limit)
    run_classifier(pairs)
